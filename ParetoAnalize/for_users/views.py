from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .forms import SaleForm, ShopForm, ShopSelectionForm, ExcelUploadForm
from openpyxl import load_workbook
from for_admin.models import MarketPlace, ModelWork, MethodofExport, TariffTranslate, CategoryYM, CategoryOZ
from .models import Shop, Sale, Product
from .until.algoritm_cattegory import SaleAnalizeCategory, ProductAnalizeCategory
from .until.algoritme_product import SaleAnalizeProduct, ProductAnalizeProduct


def main(request):
    return render(request, 'for_users/main.html')

@login_required
def profil(request):
    user = request.user

    if request.method == 'POST':
        form = ShopForm(request.POST)
        if form.is_valid():
            shop = form.save(commit=False)
            shop.creator = request.user
            shop.save()
            return redirect('profil')
    else:
        form = ShopForm()

    shops = Shop.objects.filter(creator=user)

    return render(request, 'for_users/profil.html', {'form': form, 'shops':shops})


def product(request):
    user = request.user

    if request.method == 'POST':

        form = ShopSelectionForm(user, request.POST)
        if form.is_valid():
            selected_shop = form.cleaned_data['shop']

            return redirect('load_shop_data', shop_id=selected_shop.id)
    else:
        form = ShopSelectionForm(user)

    return render(request, 'for_users/product.html', {'form': form})


def load_shop_data(request, shop_id):
    selected_shop = Shop.objects.get(id=shop_id)
    sale_pareto = []
    sale_product_pareto= []
    product_pareto = []
    product_product_pareto = []

    if request.method == 'POST':
        form_product = ExcelUploadForm(request.POST, request.FILES)
        form_sale = ExcelUploadForm(request.POST, request.FILES)
        if 'form_type' in request.POST:
            form_type = request.POST['form_type']
            if form_type == 'sale':
                if form_sale.is_valid():

                    if selected_shop.mp.nameMP == 'Яндекс.Маркет':

                        excel_file = request.FILES['excel_file']
                        wb = load_workbook(excel_file)
                        sheet = wb.worksheets[1]

                        for row in sheet.iter_rows(min_row=9, values_only=True):
                            number_shop = row[8]
                            sku_p = row[11]
                            name_p = row[12]
                            count = row[13]
                            price = row[16]

                            full_price = 0
                            count = float(count)
                            category = ''

                            tt_category = 0

                            from_product = Product.objects.all()
                            from_category = CategoryYM.objects.all()

                            for found_name in from_product:
                                if name_p == found_name.nameP:
                                    category = found_name.category
                                    full_price = found_name.price

                            for found_cost in from_category:
                                if category == found_cost.name_full:
                                    if selected_shop.md.nameMW == 'FBY':
                                        tt_category = full_price * found_cost.cost_fby
                                        print('fby')

                                    else:
                                        tt_category = full_price * found_cost.cost_fbs

                            profit = price - selected_shop.moe.cost - tt_category - price * selected_shop.tt.cost

                            sale = Sale(number_shop=number_shop, sku_p=sku_p, name_p=name_p, count=count, price=price,
                                        profit=profit, category=category, shop=selected_shop)
                            sale.save()


                    elif selected_shop.mp.nameMP == 'Озон':

                        excel_file = request.FILES['excel_file']
                        wb = load_workbook(excel_file)
                        sheet = wb.worksheets[0]

                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            number_shop = row[0]
                            sku_p = row[9]
                            name_p = row[8]
                            count = row[13]
                            price = row[11]
                            price = float(price)
                            full_price = 0
                            count = float(count)
                            category = ''

                            tt_category = 0

                            from_product = Product.objects.filter(shop=selected_shop)
                            from_category = CategoryOZ.objects.all()

                            for found_name in from_product:
                                if name_p == found_name.nameP:
                                    category = found_name.category
                                    full_price = found_name.price

                            for found_cost in from_category:
                                if category == found_cost.name_category:
                                    if selected_shop.md.nameMW == 'FBO':
                                        tt_category = full_price * found_cost.cost_fbo

                                    else:
                                        tt_category = full_price * found_cost.cost_fbs

                            profit = price - selected_shop.moe.cost - tt_category - price * selected_shop.tt.cost

                            sale = Sale(number_shop=number_shop, sku_p=sku_p, name_p=name_p, count=count, price=price,
                                        profit=profit, category=category, shop=selected_shop)
                            sale.save()

                    sale_pareto = SaleAnalizeCategory(selected_shop)
                    sale_product_pareto = SaleAnalizeProduct(selected_shop)


            if form_type == 'product':
                if form_product.is_valid():

                    if selected_shop.mp.nameMP == 'Яндекс.Маркет':

                        excel_file = request.FILES['excel_file']
                        wb = load_workbook(excel_file)
                        sheet = wb.worksheets[1]

                        for row in sheet.iter_rows(min_row=4, values_only=True):
                            sku = row[2]
                            nameP = row[3]
                            article = row[12]
                            price = row[16]
                            category = row[38]
                            category = category.replace('\\', '/')

                            tt_category = 0
                            profit = 0
                            categoryYM = CategoryYM.objects.all()

                            for category_obj in categoryYM:

                                if category == category_obj.name_full:
                                    if selected_shop.md.nameMW == 'FBY':
                                        tt_category = category_obj.cost_fby
                                    else:
                                        tt_category = category_obj.cost_fbs

                            if price != None:
                                profit = price - selected_shop.moe.cost - price * selected_shop.tt.cost - price * tt_category


                            product = Product(sku=sku, nameP=nameP, article=article, price=price, category=category, shop=selected_shop, profit=profit)
                            product.save()

                    if selected_shop.mp.nameMP == 'Озон':

                        excel_file = request.FILES['excel_file']
                        wb = load_workbook(excel_file)
                        sheet = wb.worksheets[0]

                        for row in sheet.iter_rows(min_row=4, values_only=True):
                            sku = row[9]
                            nameP = row[2]
                            article = row[1]
                            price = row[3]
                            category = row[48]

                            tt_category = 0
                            profit = 0
                            categoryOZ = CategoryOZ.objects.all()

                            for category_obj in categoryOZ:

                                if category == category_obj.name_category:
                                    if selected_shop.md.nameMW == 'FBO':
                                        tt_category = category_obj.cost_fby
                                    else:
                                        tt_category = category_obj.cost_fbs

                            if price != None:
                                profit = price - selected_shop.moe.cost - price * selected_shop.tt.cost - price * tt_category


                            product = Product(sku=sku, nameP=nameP, article=article, price=price, category=category, shop=selected_shop, profit=profit)
                            product.save()

                    product_pareto = ProductAnalizeCategory(selected_shop)
                    product_product_pareto = ProductAnalizeProduct(selected_shop)




    else:
        form_sale = ExcelUploadForm()
        form_product = ExcelUploadForm()

    return render(request, 'for_users/product.html', {'product_product_pareto': product_product_pareto, 'sale_product_pareto': sale_product_pareto, 'product_pareto': product_pareto, 'sale_pareto': sale_pareto, 'selected_shop': selected_shop, 'form_sale': form_sale, 'form_product': form_product}, )


def clients(request):
    return render(request, 'for_users/product.html')


def competitor(request):
    return render(request, 'for_users/competitor.html')


def help(request):
    return render(request, 'for_users/help.html')


