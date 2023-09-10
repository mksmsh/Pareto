from django.shortcuts import render
from .forms import ExcelUploadForm
from openpyxl import load_workbook
from .models import CategoryYM, CategoryOZ, CategorySBR


def upload_data(request):
    if request.method == 'POST':
        form_yandex = ExcelUploadForm(request.POST, request.FILES)
        form_ozone = ExcelUploadForm(request.POST, request.FILES)
        form_sbr = ExcelUploadForm(request.POST, request.FILES)

        if 'form_type' in request.POST:
            form_type = request.POST['form_type']
            if form_type == 'yandex':
                if form_yandex.is_valid():
                    excel_file = request.FILES['excel_file']
                    wb = load_workbook(excel_file)
                    sheet = wb.active

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        cell_value = row[0]
                        cell_value1 = row[1]
                        cell_value2 = row[2]
                        cell_value3 = row[3]
                        category = CategoryYM(name_parents=cell_value, name_dote=cell_value1, cost_fby=cell_value2, cost_fbs=cell_value3)
                        category.save()

            elif form_type == 'ozone':
                if form_ozone.is_valid():
                    excel_file = request.FILES['excel_file']
                    wb = load_workbook(excel_file)
                    sheet = wb.active

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        cell_value = row[0]
                        cell_value1 = row[1]
                        cell_value2 = row[2]
                        category = CategoryOZ(name_category=cell_value, cost_fbo=cell_value1, cost_fbs=cell_value2 )
                        category.save()

            elif form_type == 'sbr':
                if form_sbr.is_valid():
                    excel_file = request.FILES['excel_file']
                    wb = load_workbook(excel_file)
                    sheet = wb.active

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        cell_value = row[1]
                        cell_value1 = row[2]
                        cell_value2 = row[3]
                        category = CategorySBR(name_category=cell_value, cost_fbo=cell_value1, cost_fbs=cell_value2 )
                        category.save()

    else:
        form_yandex = ExcelUploadForm()
        form_ozone = ExcelUploadForm()
        form_sbr = ExcelUploadForm()

    return render(request, 'for_admin/up.html', {'form_yandex': form_yandex, 'form_ozone': form_ozone, 'form_sbr': form_sbr})
